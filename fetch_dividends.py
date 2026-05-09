import yfinance as yf
import pandas as pd
import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from config import TICKERS, DIVIDEND_START_DATE, DIVIDEND_END_DATE, OUTPUT_FILE


def fetch_dividends(tickers: list[str], start: str = None, end: str = None) -> pd.DataFrame:
    all_dividends = []
    for symbol in tickers:
        print(f"Fetching dividends for {symbol}...")
        ticker = yf.Ticker(symbol)
        dividends = ticker.dividends

        if dividends.empty:
            print(f"  No dividends found")
            continue

        df = pd.DataFrame({
            "Ticker": symbol,
            "Date": dividends.index,
            "Dividend": dividends.values
        })

        # Filter by date range if provided
        if start:
            df = df[df["Date"] >= start]
        if end:
            df = df[df["Date"] <= end]

        df["Date"] = df["Date"].dt.tz_localize(None)
        all_dividends.append(df)
        print(f"  {len(df)} dividends retrieved")

    if not all_dividends:
        print("No dividends found for any tickers.")
        return None

    result = pd.concat(all_dividends, ignore_index=True)
    result = result.sort_values("Date", ascending=False)
    return result


def write_to_excel(data: pd.DataFrame, output_path: str):
    # Load or create workbook
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Remove sheet if it exists
    if "Dividends" in wb.sheetnames:
        del wb["Dividends"]

    # Create new sheet and write data
    ws = wb.create_sheet("Dividends")
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Format columns
    ws.column_dimensions["A"].width = 12  # Ticker
    ws.column_dimensions["B"].width = 18  # Date
    ws.column_dimensions["C"].width = 14  # Dividend

    # Format dividend column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[2].number_format = "#,##0.00"

    wb.save(output_path)
    print(f"\nSaved to: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    data = fetch_dividends(TICKERS, start=DIVIDEND_START_DATE, end=DIVIDEND_END_DATE)

    if data is None:
        sys.exit(1)

    write_to_excel(data, OUTPUT_FILE)
